import openpyxl
import os


nurses = openpyxl.load_workbook('nurses.xlsx')['Nurses']
patients = openpyxl.load_workbook('output.xlsx')

final = openpyxl.Workbook() 
sheet = final.active

for o in range(1,len(patients.worksheets)+1):
    sheet.cell(row = o+1,column=1).value = o 
    
sheet.cell(row=1, column = 1).value = "sample"
sheet.cell(row=1, column = 2).value = "3:1 violations"
sheet.cell(row=1, column = 3).value = "1:1 violations"
sheet.cell(row=1, column = 4).value = "overrides"

count = 1
for x in patients.worksheets:
    total_pt = 0
    total_vln1 = 0 
    total_vln2 = 0
    for i in range(2,x.max_row+1):
        strt = 0
        end = 0
        cont = 0
        vln_1 = 0
        vln_2 = 0
        for j in range(3,x.max_column+1):
            strt = strt + x.cell(row=i, column=j).value 
            val = ((x.cell(row=1,column=j).value)/15) - 1
            if i - val >= 2:
                end += x.cell(row =i-val,column=j).value
            else:
                end += 0
            
            if i-val+1 < 2:
                if i == 2:
                    cont += 0
                else:
                    for k in range(2,i):
                        cont += x.cell(row=k,column=j).value
            else:
                if i == 2:
                    cont += 0
                else:
                    for k in range(int(i-val+1),i):
                        cont += x.cell(row=k,column=j).value
    
        if (strt + end + cont) <= 3*nurses.cell(row = i, column = 2).value:
            vln_1 += 0
        else:
            vln_1 += 1
        if strt + end <= nurses.cell(row = i, column = 2).value:
            vln_2 += 0
        else:
            vln_2 += 1
    
        total_vln1 += vln_1
        total_vln2 += vln_2
        

    count+= 1        
    print(total_vln1,total_vln2)    
    
    sheet.cell(row= count, column =2).value = total_vln1  
    sheet.cell(row= count, column =3).value = total_vln2

directory_in_str = "C:/Users/isikder/Pictures/Testingtesting/Current temp"
    
directory = os.fsencode(directory_in_str)
    
for file in os.listdir(directory):
    filename = os.fsdecode(file)
    if filename.endswith(".txt"):
        new = open(filename,'r')
        new = new.readlines()
        sum = 0
        for i in range(0,3):
            sum += int(new[i].split()[-1])
        for o in range(1,len(patients.worksheets)+1):
            if os.path.splitext(filename)[0] == str(o):
                sheet.cell(row= o+1, column = 4).value = sum
        print(filename, sum)    
   

final.save("C:/Users/isikder/Pictures/Testingtesting/Current temp/final current.xlsx")

         

    
    
    

    